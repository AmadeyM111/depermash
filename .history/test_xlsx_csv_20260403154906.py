#!/usr/bin/env python3
"""
Тестирование деперсонализации XLSX и CSV файлов.
"""

import tempfile
from pathlib import Path
from depersonalizer import Depersonalizer, _OPENPYXL_OK, _PANDAS_OK

def test_csv():
    """Тест деперсонализации CSV файла."""
    if not _PANDAS_OK:
        print("⚠ CSV тест пропущен: pandas не установлен")
        return

    print("=== ТЕСТ CSV ===")
    csv_content = """ФИО,Телефон,Email,Город
Иванов Иван Иванович,+7-900-123-45-67,ivanov@example.com,Москва
Петров Петр Петрович,+7-916-111-22-33,petrov@mail.ru,Санкт-Петербург
Сидорова Анна Сергеевна,+7-926-333-44-55,sidorova@gmail.com,Казань"""

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        src_csv = tmp / "test.csv"

        # Создаем тестовый файл
        src_csv.write_text(csv_content, encoding="utf-8")

        # Деперсонализируем
        dp = Depersonalizer(mode="placeholder", use_natasha=False, use_presidio=False)
        out_csv = dp.anonymize_file(src_csv)

        # Проверяем результат
        result = out_csv.read_text(encoding="utf-8")
        print("Результат:")
        print(result)

        # Проверяем, что конфиденциальные данные заменены
        assert "[ТЕЛЕФОН_" in result
        assert "[EMAIL_" in result
        assert "[ФИО_" in result

        report = dp.get_report()
        print(f"\nНайдено сущностей: {report['total_entities']}")
        print("✓ CSV тест пройден!")


def test_xlsx():
    """Тест деперсонализации XLSX файла."""
    if not _OPENPYXL_OK or not _PANDAS_OK:
        print("⚠ XLSX тест пропущен: openpyxl или pandas не установлен")
        return

    print("\n=== ТЕСТ XLSX ===")

    try:
        from openpyxl import Workbook
    except ImportError:
        print("⚠ XLSX тест пропущен: openpyxl не установлен")
        return

    with tempfile.TemporaryDirectory() as tmpdir:
        tmp = Path(tmpdir)
        src_xlsx = tmp / "test.xlsx"

        # Создаем тестовый файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Сотрудники"

        # Заголовки
        ws.append(["ФИО", "Телефон", "Email", "Паспорт"])

        # Данные
        ws.append(["Иванов Иван Иванович", "+7-900-123-45-67", "ivanov@example.com", "4510 123456"])
        ws.append(["Петров Петр Петрович", "+7-916-111-22-33", "petrov@mail.ru", "4509 654321"])
        ws.append(["Сидорова Анна Сергеевна", "+7-926-333-44-55", "sidorova@gmail.com", "4515 789012"])

        wb.save(src_xlsx)

        # Деперсонализируем
        dp = Depersonalizer(mode="placeholder", use_natasha=False, use_presidio=False)
        out_xlsx = dp.anonymize_file(src_xlsx)

        # Читаем результат
        from openpyxl import load_workbook
        wb_result = load_workbook(out_xlsx, data_only=True)
        ws_result = wb_result.active

        print("Результат:")
        for row in ws_result.iter_rows(values_only=True):
            print(f"  {row}")

        # Проверяем, что данные анонимизированы
        data = list(ws_result.iter_rows(values_only=True))
        assert any("[ТЕЛЕФОН_" in str(cell) for row in data for cell in row if cell)
        assert any("[EMAIL_" in str(cell) for row in data for cell in row if cell)
        assert any("[ФИО_" in str(cell) for row in data for cell in row if cell)

        report = dp.get_report()
        print(f"\nНайдено сущностей: {report['total_entities']}")
        print("✓ XLSX тест пройден!")


if __name__ == "__main__":
    test_csv()
    test_xlsx()
    print("\n=== ВСЕ ТЕСТЫ ПРОЙДЕНЫ ===")
